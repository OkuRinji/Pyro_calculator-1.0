from .models import Component, Library
from openpyxl import load_workbook
import psycopg2
def safe_float(x, default=0.0):
    if x is None:
        return default
    return float(x)
def load_library_from_excel(filepath:str) -> Library:
    wrkbk=load_workbook(filepath,read_only=True)
    sheet=wrkbk.active   
    element_names=[]
    new_lib=Library()
    for j in range(5,sheet.max_column+1):
            elem=sheet.cell(row=1,column=j).value
            if elem is not None:
                 element_names.append(elem)
    if len(element_names)>6:              
        comp_type=("oxy")
    else: comp_type=("fuel")
    for i in range(2, sheet.max_row+1):
        formula={}
        name = sheet.cell(row=i, column=1).value or f"Component_{i}"
        enthalpy = safe_float(sheet.cell(row=i, column=2).value) 
        demidov_coeff =safe_float(sheet.cell(row=i, column=3).value) 
        molar_mass=safe_float(sheet.cell(row=i, column=4).value) 
        for idx,j in enumerate(range(5,sheet.max_column+1)):
            value=sheet.cell(row=i, column=j).value
            element_name=element_names[idx]
            formula[element_name]=float(value) if value is not None else  0.0 

        
        new_lib.add_component(Component(type=comp_type,name=name,enthalpy=enthalpy,demidov_coeff=demidov_coeff,molar_mass=molar_mass,formula=formula))
    return new_lib


# onelist=["Mg","Ba","Sr",]
# onehalflist=["Al"]
# halflist=["K","Na"]
# Kwd={0:0.2782,1:1.084,2:1.4,3:2.3186,4:2.7984,5:3.2103,6:3.467,7:3.853,8:4.625,9:5.332,10:5.968,11:6.2365,12:6.538,13:7.436,}
#"3200":7.886,"3300":8.15,"3400":8.29,"3600":8.54